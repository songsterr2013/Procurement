import os
from xlrd import open_workbook
from openpyxl import load_workbook

import util

util.running_prerequisite()
logger = util.get_logger(__file__)
s, config = util.load_config()


class Procurement:

    excluded_type = config["exclude_type"].split(',')
    written_row_index = 1
    written_column_index = 1

    def __init__(self, main_excel, save_path, bom_path, pallet_path):

        self.main_excel = main_excel
        self.save_path = save_path
        self.bom_path = bom_path
        self.pallet_path = pallet_path

        self.wb = load_workbook(main_excel, read_only=False)

    # get data
    def read_main_excel(self):
        table = self._get_worksheet()

        for row_index, row in enumerate(table.iter_rows(min_row=2, min_col=1, max_col=10, values_only=True)):
            if row[0] is not None:
                yield row

    def write_data(self, origin, pallet, bom):
        written_table = self._get_new_worksheet()

        if len(pallet) == 0 and len(bom) == 0:
            self.write_action_1(origin)
            self.write_action_2(origin)

            self.written_row_index += 1

        elif len(pallet) == 0 and len(bom) != 0:
            self.write_action_1(origin)
            self.write_action_2(origin)
            for row in bom:

                for val in row:
                    written_table.cell(row=self.written_row_index, column=self.written_column_index).value = val
                    self.written_column_index += 1

                written_table.cell(row=self.written_row_index, column=self.written_column_index).value = row[3] * origin[8]

                self.written_column_index -= len(row)
                self.written_row_index += 1

        elif len(pallet) != 0 and len(bom) == 0:
            self.write_action_1(origin)
            for row in pallet:
                for val in row:
                    written_table.cell(row=self.written_row_index, column=self.written_column_index).value = val
                    self.written_column_index += 1

                self.written_column_index -= len(row)
                self.written_row_index += 1

        elif len(pallet) != 0 and len(bom) != 0:
            self.write_action_1(origin)
            for P_row in pallet:

                written_table.cell(row=self.written_row_index, column=self.written_column_index).value = P_row[0]
                self.written_column_index += 1
                written_table.cell(row=self.written_row_index, column=self.written_column_index).value = P_row[1]
                self.written_column_index += 1

                for B_row in bom:
                    written_table.cell(row=self.written_row_index, column=self.written_column_index).value = P_row[2]
                    self.written_column_index += 2

                    for B_val in B_row:
                        written_table.cell(row=self.written_row_index, column=self.written_column_index).value = B_val
                        self.written_column_index += 1

                    written_table.cell(row=self.written_row_index, column=self.written_column_index).value = B_row[3] * P_row[1]

                    self.written_column_index -= len(B_row) + 2
                    self.written_row_index += 1

                self.written_column_index -= 2

        self.written_column_index = 1

    def write_action_1(self, origin):
        written_table = self._get_new_worksheet()

        for val in origin:
            written_table.cell(row=self.written_row_index, column=self.written_column_index).value = val
            self.written_column_index += 1

    def write_action_2(self, origin):
        written_table = self._get_new_worksheet()

        written_table.cell(row=self.written_row_index, column=self.written_column_index).value = origin[6]
        self.written_column_index += 1
        written_table.cell(row=self.written_row_index, column=self.written_column_index).value = origin[8]
        self.written_column_index += 3

    def get_required_pallet_data(self, order_name, file_name):
        product_name_loc = 0
        amount_loc = 0
        dispatch_order_loc = 0
        tmp_dispatch_order = ''

        for index, row in enumerate(self.yield_pallet_content(order_name)):
            # 取得column位置
            if index == 0:
                for index_2, row_2 in enumerate(row):
                    if row_2 == '產品編號':
                        product_name_loc = index_2
                    elif row_2 == '數量':
                        amount_loc = index_2
                    elif row_2 == '派工單':
                        dispatch_order_loc = index_2

            # 合併儲存格的關係
            if row[dispatch_order_loc] is not None:
                tmp_dispatch_order = row[dispatch_order_loc]

            if row[product_name_loc] == file_name:
                yield row[product_name_loc], row[amount_loc], tmp_dispatch_order

    def yield_pallet_content(self, order_name):
        path = self.pallet_path
        completed_pallet_path = os.path.join(path, order_name)
        if os.path.isfile(completed_pallet_path + '.xls'):
            target_wb = open_workbook(completed_pallet_path + '.xls')
            target_sh = target_wb.sheets()[0]
            for row in range(0, target_sh.nrows):
                if type(target_sh.row_values(row)[0]) != str:
                    yield target_sh.row_values(row)

        elif os.path.isfile(completed_pallet_path + '.xlsx'):
            target_wb = load_workbook(completed_pallet_path + '.xlsx', read_only=False)
            target_sh = target_wb[target_wb.sheetnames[0]]
            for row in target_sh.iter_rows(min_row=1, min_col=1, max_col=12, values_only=True):
                if row[0] is not None:
                    yield row

    # get specific data
    def get_required_bom_data(self, file_name):

        for row in self.yield_bom_content(file_name):
            if row[7] == '' or row[7] is None:
                raise ValueError('{}格式異常'.format(file_name))
            elif row[7] not in self.excluded_type:
                if str(row[7]) != 'SPC':
                    yield row[7], row[1], row[3], int(row[4])
                elif str(row[7]) == 'SPC' and float(row[3].split('\\')[0][:-1]) >= 12.0:
                    yield row[7], row[1], row[3], int(row[4])

    # decide format and yield content
    def yield_bom_content(self, file_name):
        path = self.bom_path
        prefix = file_name[0]
        completed_path = os.path.join(path, prefix, file_name)

        if os.path.isfile(completed_path + '.xls'):
            target_wb = open_workbook(completed_path + '.xls')
            target_sh = target_wb.sheets()[0]
            for row in range(2, target_sh.nrows):
                if type(target_sh.row_values(row)[0]) != str:
                    yield target_sh.row_values(row)
        elif os.path.isfile(completed_path + '.xlsx'):
            target_wb = load_workbook(completed_path + '.xlsx', read_only=False)
            target_sh = target_wb[target_wb.sheetnames[0]]
            for row in target_sh.iter_rows(min_row=3, min_col=1, max_col=12, values_only=True):
                if row[0] is not None:
                    yield row
        # xls或xlsx都找不到的話就只能在LOG顯示，但EXCEL上會保留此列:
        else:
            logger.info("沒有該BOM檔案:{}".format(file_name))

    def _get_worksheet(self):
        return self.wb[self.wb.sheetnames[0]]

    def make_new_sheet(self):
        new_ws = self.wb.create_sheet('採購輔助程式結果')

        new_ws.cell(row=self.written_row_index, column=1, value="預交貨日")
        new_ws.cell(row=self.written_row_index, column=2, value="客戶簡稱")
        new_ws.cell(row=self.written_row_index, column=3, value="單據號碼")
        new_ws.cell(row=self.written_row_index, column=4, value="專案名稱")
        new_ws.cell(row=self.written_row_index, column=5, value="機號")
        new_ws.cell(row=self.written_row_index, column=6, value="欄號")
        new_ws.cell(row=self.written_row_index, column=7, value="產品編號")
        new_ws.cell(row=self.written_row_index, column=8, value="品名規格")
        new_ws.cell(row=self.written_row_index, column=9, value="數量")
        new_ws.cell(row=self.written_row_index, column=10, value="製單人員")
        new_ws.cell(row=self.written_row_index, column=11, value="產品編號")
        new_ws.cell(row=self.written_row_index, column=12, value="數量")
        new_ws.cell(row=self.written_row_index, column=13, value="派工單/棧板數")
        new_ws.cell(row=self.written_row_index, column=14, value="採購日")
        new_ws.cell(row=self.written_row_index, column=15, value="類別")
        new_ws.cell(row=self.written_row_index, column=16, value="料號")
        new_ws.cell(row=self.written_row_index, column=17, value="品名規格")
        new_ws.cell(row=self.written_row_index, column=18, value="數量")
        new_ws.cell(row=self.written_row_index, column=19, value="採購量")
        new_ws.cell(row=self.written_row_index, column=20, value="採單單號")
        new_ws.cell(row=self.written_row_index, column=21, value="備註")

        self.written_row_index += 1

    def _get_new_worksheet(self):
        return self.wb['採購輔助程式結果']

    def save(self):
        self.wb.save(self.save_path)
