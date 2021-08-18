import json
import os
import sys
import logging
from logging.handlers import TimedRotatingFileHandler
from logging import StreamHandler
from xlrd import open_workbook
from openpyxl import load_workbook


def get_logger(name):
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    # 這是把error traceback寫到log中的一個module，配合traceback，
    # 原先traceback因為被raise error了他不會被秀出來
    handler = TimedRotatingFileHandler("log/api_util.log",
                                       encoding="UTF-8",
                                       when="d",
                                       interval=1,
                                       backupCount=7)

    # 簡單來說這個東西是另一種把error print出來的一個module
    std_handler = StreamHandler(sys.stdout)

    formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s', datefmt='%m-%d %H:%M:%S')
    handler.setFormatter(formatter)
    std_handler.setFormatter(formatter)

    logger.addHandler(handler)
    logger.addHandler(std_handler)

    return logger



'''def get_bom(bom_path, file_name):
    path = bom_path
    prefix = file_name[0]

    completed_path = os.path.join(path, prefix, file_name)

    if os.path.isfile(completed_path + '.xls'):
        target_wb = open_workbook(completed_path + '.xls')
        target_sh = target_wb.sheets()[0]
        return 'xls', target_sh
    elif os.path.isfile(completed_path + '.xlsx'):
        target_wb = load_workbook(completed_path + '.xlsx', read_only=False)
        target_sh = target_wb[target_wb.sheetnames[0]]
        return 'xlsx', target_sh
    # xls或xlsx都找不到的話就只能:
    else:
        return False, None'''


def load_config():
    root_path = ""
    if getattr(sys, 'frozen', False):
        root_path = os.path.dirname(sys.executable)
    elif __file__:
        root_path = os.path.dirname(__file__)

    config_file_path = os.path.join(root_path, "config.json")
    if os.path.exists(config_file_path):
        with open(config_file_path, 'r', encoding="utf-8") as fp:
            config = json.load(fp)
        return True, config
    else:
        return False, None


def running_prerequisite():
    root_path = ''

    if getattr(sys, 'frozen', False):  # 如果為exe
        root_path = os.path.dirname(sys.executable)
    elif __file__:
        root_path = os.path.dirname(__file__)

    if not os.path.isdir(os.path.join(root_path, "log")):
        os.mkdir(os.path.join(root_path, "log"))
